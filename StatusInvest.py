import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By

# define selenium webdriver options
options = webdriver.ChromeOptions()

import warnings
url = 'https://fundamentus.com.br/detalhes.php?papel=DASA3&interface=classic&interface=mobile'
# create selenium webdriver instance
driver = webdriver.Chrome(options=options)
warnings.filterwarnings('ignore')
headers = {
    'User-Agent'      : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
    'Accept'          : 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language' : 'en-US,en;q=0.5',
    'DNT'             : '1',
    'Connection'      : 'close'
}
wbsaida = openpyxl.Workbook()

#cria planilhas
#cria planilhas
def criaPlanilaIndValuation(wbsaida):
    wbsaida.create_sheet('IndValuation')
    IndValuation = wbsaida['IndValuation']
    IndValuation.append(['D.Y', 'P/L', ' PEG Ratio','P/VP','EV/EBITDA','EV/EBIT','P/EBITDA','P/EBIT','VPA','P/Ativo',
                         'LPA','P/SR','P/Ativo Circ. Liq.'])
    return


def criaPlanilhaIndEndividamento(IndEndividamento):
    wbsaida.create_sheet('IndEndividamento')
    IndEndividamento = wbsaida['IndEndividamento']
    IndEndividamento.append(['Dív. líquida/PL', 'Dív. líquida/EBITDA', 'Dív. líquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente'])
    return

def criaPlanilhaIndiEficiência(IndiEficiência):
    wbsaida.create_sheet('IndiEficiência')
    IndiEficiência = wbsaida['IndiEficiência']
    IndiEficiência.append(['M. Bruta', 'M. EBITDA', 'M. EBIT', 'M. Líquida'])
    return

def criaPlanilhaIndRentabilidade(IndiRentabilidade):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(['ROE', 'ROA', 'ROIC','Giro ativos',''])
    return
def criaPlanilhaIndiCrescimento(IndiCrescimento):
    wbsaida.create_sheet('IndiCrescimento')
    IndiCrescimento = wbsaida['IndiCrescimento']
    IndiCrescimento.append(['CAGR Receitas 5 anos', 'CAGR Lucros 5 anos'])
    return
def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock url
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []
    data = requests.get(url, headers=headers, timeout=6).text
    soup10 = BeautifulSoup(data, "html.parser")

    # get divs from stock <h1 class="lh-4" title="ATOM3 - ATOM EMPREENDIMENTOS">ATOM3 - <small>ATOM EMPREENDIMENTOS</small></h1>
    company_name = soup10.find('span', {'class': 'acao-nome'}).text
    print ("vary  " + company_name)
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]
    teste = soup2.find_all('strong','value')

   # print(teste)
    #print(soups)


    for s in soups:
       # print(s)
        # get only titles from a div and append to keys
        #print("re   " + re.compile('title m-0[^"]*'))
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        #titles = s.find_all('h3', re.compile('title m-0'))
        titles2 = s.find_all('h3', 'title m-0')
        #print(titles2)
        #print(titles)
        titles = [t.get_text() for t in titles]
        #print(titles)
        keys += titles
       # print(keys)
        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        #print(numbers)
        numbers = [n.get_text()for n in numbers]
        #print(numbers)
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}
    #print(d)
    return d

def gravaIndiRentabilidade(wsIndiRentabilidade,teste,coluna,valor):
    print(" teste  "  + valor)
    wsIndiRentabilidade.cell(row=teste, column=coluna, value=valor)

    return

if __name__ == "__main__":
    dict_stocks = {}

    # start timer
    start = time.time()
    criaPlanilaIndValuation(wbsaida)
    criaPlanilhaIndEndividamento(wbsaida)
    criaPlanilhaIndiEficiência(wbsaida)
    criaPlanilhaIndRentabilidade(wbsaida)
    criaPlanilhaIndiCrescimento(wbsaida)

    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()
        linha = 1
        # get stock information and create excel sheet
        for stock in stocks:

            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                print(stock)
                print(dict_stocks[stock].get("Liq. corrente"))
                print(dict_stocks[stock].get("M. Bruta"))
                wsIndiRentabilidade = wbsaida['IndiRentabilidade']
                linha = linha + 1
                gravaIndiRentabilidade(wsIndiRentabilidade, linha,2, dict_stocks[stock].get("Liq. corrente"))
                #gravaIndiRentabilidade(wsIndiRentabilidade,teste, 2, stock)


            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information')

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("exemplo2.xlsx")
    print(f'Brasilian stocks information got in {int(end-start)} s')



