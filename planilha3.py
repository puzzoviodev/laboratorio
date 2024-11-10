import openpyxl

wb = openpyxl.load_workbook('statusinvest2.xlsx')

ws = wb['st']
#ws2 = wb['silvio']


#IndValuation
#IndEndividamento
#IndiEficiência
#IndiRentabilidade
#IndiCrescimento

#print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

#print('The value in cell A1 is: '+ws['A1'].value)

#values = [ws.cell(row=i,column=i).value for i in range(10,ws.max_row)]
#print(values)
# Inicializa listas para armazenar os dados

def Myfun(imposto):
    values2 = ws.cell(row=imposto, column=2).value
    print("func " + values2)
    return

imposto = 1
#while imposto < ws.max_row:
while imposto < 10:
    imposto = imposto + 1

   # print(imposto)
    values = ws.cell(row=imposto, column=1).value
    print(" valoe  "  + values)
    print(imposto)
    var = 'K' + str(imposto)
    var1= 'A' + str(imposto)
   # print(var)
    ws[var] = values
    #ws2[var1] = values
   # retorno1 = Myfun(imposto)
#
#wb.save('videogamesales.xlsx')
