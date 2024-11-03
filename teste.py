#IndValuation
#IndEndividamento
#IndiEficiência
#IndiRentabilidade
#IndiCrescimento


import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# Criar uma nova planilha
wb = openpyxl.Workbook()
sheet = wb.active

wb = openpyxl.load_workbook('statusinvest2.xlsx')
book = openpyxl.Workbook()

ws = wb['statusinvest']



import openpyxl

# Criar uma nova planilha


# Criar uma nova aba e adicionar dados nela
book.create_sheet('IndValuation')
book.create_sheet('IndEndividamento')
book.create_sheet('IndiEficiência')
book.create_sheet('IndiRentabilidade')
book.create_sheet('IndiCrescimento')


IndValuation  =  book['IndValuation']
IndEndividamento  = book['IndEndividamento']
IndiEficiência  = book['IndiEficiência']
IndiRentabilidade =  book['IndiRentabilidade']
IndiCrescimento = book['IndiCrescimento']


IndValuation.append(['Fruta', 'Quantidade', 'Preço'])
IndEndividamento.append(['Fruta', 'Quantidade', 'Preço'])
IndiEficiência.append(['Fruta', 'Quantidade', 'Preço'])
IndiRentabilidade.append(['Fruta', 'Quantidade', 'Preço'])
IndiCrescimento.append(['Fruta', 'Quantidade', 'Preço'])



ws2 = book['IndValuation']

imposto = 1
#while imposto < ws.max_row:
while imposto < 10:
    imposto = imposto + 1

   # print(imposto)

    values = ws.cell(row=imposto, column=1).value
    values2 = ws.cell(row=imposto, column=2).value

    print(imposto)
    var1= 'A' + str(imposto)
    var2 = 'B' + str(imposto)
    ws2[var1] = values
    ws2[var2] = values2
       # print(var)




# Salvar a planilha
book.save('minha_planilha7.xlsx')



