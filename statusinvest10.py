import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import warnings


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instance
driver = webdriver.Chrome(options=options)

#silvio
wbsaida = openpyxl.Workbook()

#Silvio inicio
def criaPlanilaIndValuation(wbsaida):
    wbsaida.create_sheet('IndValuation')
    IndValuation = wbsaida['IndValuation']
    IndValuation.append(['D.Y', 'P/L', ' PEG Ratio','P/VP','EV/EBITDA','EV/EBIT','P/EBITDA','P/EBIT','VPA','P/Ativo',
                         'LPA','P/SR','P/Ativo Circ. Liq.'])
    return


def criaPlanilhaIndEndividamento(IndEndividamento):
    wbsaida.create_sheet('IndEndividamento')
    IndEndividamento = wbsaida['IndEndividamento']
    IndEndividamento.append(['Dív. líquida/PL', 'Dív. líquida/EBITDA', 'Dív. líquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente'])
    return

def criaPlanilhaIndiEficiência(IndiEficiência):
    wbsaida.create_sheet('IndiEficiência')
    IndiEficiência = wbsaida['IndiEficiência']
    IndiEficiência.append(['M. Bruta', 'M. EBITDA', 'M. EBIT', 'M. Líquida'])
    return

def criaPlanilhaIndRentabilidade(IndiRentabilidade):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(['ROE', 'ROA', 'ROIC','Giro ativos',''])
    return
def criaPlanilhaIndiCrescimento(IndiCrescimento):
    wbsaida.create_sheet('IndiCrescimento')
    IndiCrescimento = wbsaida['IndiCrescimento']
    IndiCrescimento.append(['CAGR Receitas 5 anos', 'CAGR Lucros 5 anos'])
    return

def gravaIndiRentabilidade(wsIndiRentabilidade,linha,coluna,valor):
        wsIndiRentabilidade.cell(row=linha, column=coluna, value=valor)
        wsIndiRentabilidade.cell(row=linha, column=coluna, value=valor)
        wsIndiRentabilidade.cell(row=linha, column=coluna, value=valor)
#Silvio fim

def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock url
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}

    # start timer
    start = time.time()
    # Silvio Inicio
    criaPlanilaIndValuation(wbsaida)
    criaPlanilhaIndEndividamento(wbsaida)
    criaPlanilhaIndiEficiência(wbsaida)
    criaPlanilhaIndRentabilidade(wbsaida)
    criaPlanilhaIndiCrescimento(wbsaida)
    # Silvio Fim
    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()
        linha = 1 # silvio
        # get stock information and create excel sheet
        for stock in stocks:
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                linha = linha + 1 #silvio

                #IndiRentabilidade
                print(dict_stocks[stock].get("ROE"))
                print(dict_stocks[stock].get("ROA"))
                print(dict_stocks[stock].get("ROIC"))
                print(dict_stocks[stock].get("Giro ativos"))
                ROE =dict_stocks[stock].get("ROE")
                ROA =dict_stocks[stock].get("ROA")
                ROIC =dict_stocks[stock].get("ROIC")
                Giroativos =dict_stocks[stock].get("Giro ativos")


                wsIndiRentabilidade = wbsaida['IndiRentabilidade']


                gravaIndiRentabilidade(wsIndiRentabilidade, linha,1, ROE)
                gravaIndiRentabilidade(wsIndiRentabilidade, linha, 2, ROA)
                gravaIndiRentabilidade(wsIndiRentabilidade, linha, 3, ROIC)
                gravaIndiRentabilidade(wsIndiRentabilidade, linha, 4, Giroativos)
                #IndiCrescimento

                print(dict_stocks[stock].get("CAGR Receitas 5 anos"))
                print(dict_stocks[stock].get("CAGR Lucros 5 anos"))

                #IndiEficiência
                print(dict_stocks[stock].get("M. Bruta"))
                print(dict_stocks[stock].get("M. EBITDA"))
                print(dict_stocks[stock].get("M. EBIT"))
                print(dict_stocks[stock].get("M. Liquida"))

                #IndEndividamento
                print(dict_stocks[stock].get("Div. liquida/PL"))
                print(dict_stocks[stock].get("Div. liquida/EBITDA"))
                print(dict_stocks[stock].get("Div. liquida/EBIT"))
                print(dict_stocks[stock].get("PL/Ativos"))
                print(dict_stocks[stock].get("Passivos/Ativos"))
                print(dict_stocks[stock].get("Liq. corrente"))

               # IndValuation
                print(dict_stocks[stock].get("D.Y"))
                print(dict_stocks[stock].get("P/L"))
                print(dict_stocks[stock].get("PEG Ratio"))
                print(dict_stocks[stock].get("P/VP"))
                print(dict_stocks[stock].get("EV/EBITDA"))
                print(dict_stocks[stock].get("EV/EBIT"))
                print(dict_stocks[stock].get("P/EBITDA"))
                print(dict_stocks[stock].get("P/EBIT"))
                print(dict_stocks[stock].get("VPA"))
                print(dict_stocks[stock].get("P/Ativo"))
                print(dict_stocks[stock].get("LPA"))
                print(dict_stocks[stock].get("P/SR"))
                print(dict_stocks[stock].get("P/Cap. Giro"))
                print(dict_stocks[stock].get("P/Ativo Circ. Liq."))

                # Diversas

                print(dict_stocks[stock].get("Valor atual"))
                print(dict_stocks[stock].get("Min. 52 semanas"))
                print(dict_stocks[stock].get("Max. 52 semanas"))
                print(dict_stocks[stock].get("dividend Yield"))
                print(dict_stocks[stock].get("Valorizacao (12m)"))
                print(dict_stocks[stock].get("Tipo"))
                print(dict_stocks[stock].get("TAG ALONG"))
                print(dict_stocks[stock].get("LIQUIDEZ MEDIA DIARIA"))
                print(dict_stocks[stock].get("PARTICIPACAO NO IBOV"))
                print(dict_stocks[stock].get("MERCADO DE OPCOES"))





                print(dict_stocks[stock].get("Patrimonio liquido"))
                print(dict_stocks[stock].get("Ativos"))
                print(dict_stocks[stock].get("Ativo circulante"))
                print(dict_stocks[stock].get("Divida bruta"))
                print(dict_stocks[stock].get("Disponibilidade"))
                print(dict_stocks[stock].get("Divida liquida"))
                print(dict_stocks[stock].get("Valor de mercado"))
                print(dict_stocks[stock].get("Valor de firma"))

            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information')

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("exemplo2.xlsx") # silvio
    print(f'Brasilian stocks information got in {int(end-start)} s')