import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# Criar uma nova planilha
workbook = openpyxl.Workbook()
sheet = workbook.active

# Adicionar dados
sheet['A1'] = 'Produto'
sheet['B1'] = 'Preço'

sheet['A2'] = 'Produto A'
sheet['B2'] = 150

sheet['A3'] = 'Produto B'
sheet['B3'] = 80

# Formatar cabeçalho
sheet['A1'].font = Font(bold=True)
sheet['B1'].font = Font(bold=True)

# Preencher célula A1 com amarelo
sheet['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Formatar B2 como moeda
sheet['B2'].number_format = 'R$ #,##0.00'
sheet['B3'].number_format = 'R$ #,##0.00'

# Adicionar formatação condicional para preços acima de 100
rule = Rule(type="expression", dxf=DifferentialStyle(fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")))
rule.formula = ["$B2>100"]
sheet.conditional_formatting.add("B2:B100", rule)

# Salvar o arquivo
workbook.save('minha_planilha_formatada.xlsx')

import openpyxl
from openpyxl import Workbook

# Criar uma nova planilha
wb = Workbook()
ws = wb.active

# Definir cabeçalhos
ws['A1'] = 'Nome'
ws['B1'] = 'Nota'
ws['C1'] = 'Porcentagem'

# Adicionar dados
dados = [
    ('Alice', 85),
    ('Bob', 92),
    ('Charlie', 78),
    ('Diana', 88)
]

# Preencher os dados e formatar a coluna de porcentagem
for i, (nome, nota) in enumerate(dados, start=2):
    ws.cell(row=i, column=1, value=nome)  # Nome do aluno
    ws.cell(row=i, column=2, value=nota)  # Nota do aluno

    # Calcular a porcentagem (por exemplo, nota / 100)
    porcentagem = nota / 100
    porcentagem_cell = ws.cell(row=i, column=3, value=porcentagem)

    # Definir o formato de porcentagem
    porcentagem_cell.number_format = '0%'  # Formato de porcentagem

# Salvar a planilha
wb.save("notas_alunos.xlsx")

