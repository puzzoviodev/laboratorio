import openpyxl
import requests

from bs4 import BeautifulSoup

import warnings

warnings.filterwarnings('ignore')


#simulação de conexão comum evitando que sistema do site entenda que é um bot bloqueando o acesso
headers = {
    'User-Agent'      : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
    'Accept'          : 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language' : 'en-US,en;q=0.5',
    'DNT'             : '1',
    'Connection'      : 'close'
}
wb = openpyxl.load_workbook('statusinvest2.xlsx')

ws = wb['st']
#ws2 = wb['silvio']

#url = 'https://fundamentus.com.br/detalhes.php?papel=DASA3&interface=classic&interface=mobile'


#data = requests.get(url, headers=headers, timeout=6).text

#coletando o html
#soup = BeautifulSoup(data, "html.parser")


imposto = 1
while imposto < ws.max_row:

      try:
            imposto = imposto + 1

            print(imposto)
            values = ws.cell(row=imposto, column=1).value

            #print(" valoe  "  + values)
            url1 = 'https://fundamentus.com.br/detalhes.php?papel='
            url2 = str(values) #'APTI3
            url3 = '&interface=classic&interface=mobile'

            url = url1 + url2 + url3
            #print('URL  ' + str(url))

            data = requests.get(url, headers=headers, timeout=6).text

            # coletando o html
            soup = BeautifulSoup(data, "html.parser")
            ticket_symbol = soup.find('h1', {'class': 'acao-papel'}).text



            company_name = soup.find('span', {'class': 'acao-nome'}).text
            frame_cotacao = soup.find('div', {'class': 'frame-cotacao'})

           print('company' + company_name)
      except AttributeError:
            print('erro')
            #ticket_symbol = "teste"
            continue
#wb.save('videogamesales.xlsx')
