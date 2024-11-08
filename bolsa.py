import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

#url do site que será feita a coleta dos dados
url = 'https://www.fundamentus.com.br/resultado.php'

#simulação de conexão comum evitando que sistema do site entenda que é um bot bloqueando o acesso
headers = {
    'User-Agent'      : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
    'Accept'          : 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language' : 'en-US,en;q=0.5',
    'DNT'             : '1',
    'Connection'      : 'close'
}




#IndValuation
#IndEndividamento
#IndiEficiência
#IndiRentabilidade
#IndiCrescimento
#wb = openpyxl.load_workbook('statusinvest2.xlsx')
wbsaida = openpyxl.Workbook()
#ws = wbsaida.active

redFill = PatternFill(start_color='FFEE1111',
end_color='FFEE1111',
fill_type='solid')


#cria planilhas
def criaPlanilaIndValuation(wbsaida):
    wbsaida.create_sheet('IndValuation')
    IndValuation = wbsaida['IndValuation']
    IndValuation.append(['D.Y', 'P/L', ' PEG Ratio','P/VP','EV/EBITDA','EV/EBIT','P/EBITDA','P/EBIT','VPA','P/Ativo',
                         'LPA','P/SR','P/Ativo Circ. Liq.'])
    return


def criaPCelulaIndEndividamento(IndEndividamento):
    wbsaida.create_sheet('IndEndividamento')
    IndEndividamento = wbsaida['IndEndividamento']
    IndEndividamento.append(['Dív. líquida/PL', 'Dív. líquida/EBITDA', 'Dív. líquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente'])
    return

def criaCelulaIndiEficiência(IndiEficiência):
    wbsaida.create_sheet('IndiEficiência')
    IndiEficiência = wbsaida['IndiEficiência']
    IndiEficiência.append(['M. Bruta', 'M. EBITDA', 'M. EBIT', 'M. Líquida'])
    return

def criaCelulaIndiRentabilidade(IndiRentabilidade):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(['ROE', 'ROA', 'ROIC','Giro ativos',''])
    return
def criaCelulaIndiCrescimento(IndiCrescimento):
    wbsaida.create_sheet('IndiCrescimento')
    IndiCrescimento = wbsaida['IndiCrescimento']
    IndiCrescimento.append(['CAGR Receitas 5 anos', 'CAGR Lucros 5 anos'])
    return

def gravaCelulaIndiRentabilidade(imposto,wsIndiRentabilidade,te1):

    #values = ws.cell(row=imposto, column=1).value
   # var1 = 'A' + str(imposto)
    nota = imposto + te1
    porcentagem = nota / 100
    #celula[var1] = porcentagem

    #porcentagem_cell = celula[var1] # ws.cell(row=i, column=3, value=porcentagem) b
    #porcentagem_cell =  celula.cell(row=imposto, column=1, value=porcentagem)
    # Definir o formato de porcentagem
    #porcentagem_cell.number_format = '0%'  # Formato de porcentagem
    wsIndiRentabilidade.cell(row=imposto, column=1, value=porcentagem).number_format = '0%'  # Formato de porcentagem
    return

criaPlanilaIndValuation(wbsaida)
criaPCelulaIndEndividamento(wbsaida)
criaCelulaIndiEficiência(wbsaida)
criaCelulaIndiRentabilidade(wbsaida)
criaCelulaIndiCrescimento(wbsaida)

linha = 1
while linha < 10:
    linha = linha + 1
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']
    wsIndValuation = wbsaida['IndValuation']
    wsIndEndividamento = wbsaida['Endividamento']
    wsIndiEficiência = wbsaida['IndiEficiência']
    wsIndiCrescimento = wbsaida['IndiCrescimento']

    gravaCelulaIndiRentabilidade(linha,wsIndiRentabilidade,1)

    print(linha)

wbsaida.save("exemplo2.xlsx")
