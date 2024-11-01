import openpyxl
import os


def Myfun(x,y=4,prompt=True):
    res = x*y
    if prompt:
        print("resposta é %i", res)

    return res
current_directory = os.getcwd()
print("O diretório de trabalho atual é:", current_directory)

#import openpyxl
import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 42
ws.title = "First22"

ws2 = wb.create_sheet()
ws2.title = "Second sheet"
ws2['A1'] = datetime.datetime.now()
ws2.sheet_properties.tabColor = "1072BA"

wb.save("two_worksheets2.xlsx")



#= 1

#print(i)
#while i <= 100:
#    conta2 = 'c' + str(i)

#    faker = Faker()
#    nome = faker.name()
#    print(nome)
    # numero = random.randint(1, 100)
#    sheet[conta2] = nome
#    i += 1

# Salve o arquivo do Excel
#workbook.save("D:/Acer 1430/Projetos Python/pythonprojeto3/exemplo29.xlsx")
print("55")

