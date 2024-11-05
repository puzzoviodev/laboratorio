import openpyxl
from openpyxl import Workbook

# Criar uma nova planilha
wb = Workbook()
ws = wb.active

# Definir cabeçalhos
ws['A1'] = 'Nome'
ws['B1'] = 'Nota'
ws['C1'] = 'Porcentagem'

# Adicionar dados
dados = [
    ('Alice', 85),
    ('Bob', 92),
    ('Charlie', 78),
    ('Diana', 88)
]

# Preencher os dados e formatar a coluna de porcentagem
for i, (nome, nota) in enumerate(dados, start=2):
    ws.cell(row=i, column=1, value=nome)  # Nome do aluno
    ws.cell(row=i, column=2, value=nota)  # Nota do aluno

    # Calcular a porcentagem (por exemplo, nota / 100)
    porcentagem = nota / 100
    porcentagem_cell = ws.cell(row=i, column=3, value=porcentagem)

    # Definir o formato de porcentagem
    porcentagem_cell.number_format = '0%'  # Formato de porcentagem

# Salvar a planilha
wb.save("notas_alunos.xlsx")

