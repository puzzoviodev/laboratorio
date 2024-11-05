from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

wb = Workbook()
ws = wb.active

# Create fill
redFill = PatternFill(start_color='FFEE1111',
end_color='FFEE1111',
fill_type='solid')


ws.conditional_formatting.add('A1:A10',
ColorScaleRule(start_type='min', start_color='FFAA0000',
end_type='max', end_color='FF00AA00')
)

ws.conditional_formatting.add('B1:B10',
ColorScaleRule(start_type='percentile', start_value=10, start_color='FFAA0000',
mid_type='percentile', mid_value=50, mid_color='FF0000AA',
end_type='percentile', end_value=90, end_color='FF00AA00')
 )

ws.conditional_formatting.add('C2:C10',
CellIsRule(operator='lessThan', formula=['C$1'], stopIfTrue=True, fill=redFill))

ws.conditional_formatting.add('D2:D10',
CellIsRule(operator='between', formula=['1','5'], stopIfTrue=True, fill=redFill))

ws.conditional_formatting.add('E1:E10',
FormulaRule(formula=['ISBLANK(E1)'], stopIfTrue=True, fill=redFill))

myFont = Font()
myBorder = Border()
ws.conditional_formatting.add('E1:E10',
FormulaRule(formula=['E1=0'], font=myFont, border=myBorder, fill=redFill))

wb.save("test.xlsx")