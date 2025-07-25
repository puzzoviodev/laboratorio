import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests

from bs4 import BeautifulSoup
# teste
import warnings
warnings.filterwarnings('ignore')





#IndValuation
#IndEndividamento
#IndiEficiência
#IndiRentabilidade
#IndiCrescimento
#wb = openpyxl.load_workbook('statusinvest2.xlsx')
wbsaida = openpyxl.Workbook()
#ws = wbsaida.active
url = 'https://www.fundamentus.com.br/resultado.php'
redFill = PatternFill(start_color='FFEE1111',
end_color='FFEE1111',
fill_type='solid')


headers = {
    'User-Agent'      : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
    'Accept'          : 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language' : 'en-US,en;q=0.5',
    'DNT'             : '1',
    'Connection'      : 'close'
}

data = requests.get(url, headers=headers, timeout=6).text

#coletando o html
soup = BeautifulSoup(data, "html.parser")

#atribuindo a variável tabela
tabela = soup.find('table')


#cria planilhas
def criaPlanilaIndValuation(wbsaida):
    wbsaida.create_sheet('IndValuation')
    IndValuation = wbsaida['IndValuation']
    IndValuation.append(['D.Y', 'P/L', ' PEG Ratio','P/VP','EV/EBITDA','EV/EBIT','P/EBITDA','P/EBIT','VPA','P/Ativo',
                         'LPA','P/SR','P/Ativo Circ. Liq.'])
    return


def criaPlanilhaIndEndividamento(IndEndividamento):
    wbsaida.create_sheet('IndEndividamento')
    IndEndividamento = wbsaida['IndEndividamento']
    IndEndividamento.append(['Dív. líquida/PL', 'Dív. líquida/EBITDA', 'Dív. líquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente'])
    return

def criaPlanilhaIndiEficiência(IndiEficiência):
    wbsaida.create_sheet('IndiEficiência')
    IndiEficiência = wbsaida['IndiEficiência']
    IndiEficiência.append(['M. Bruta', 'M. EBITDA', 'M. EBIT', 'M. Líquida'])
    return

def criaPlanilhaIndRentabilidade(IndiRentabilidade):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(['ROE', 'ROA', 'ROIC','Giro ativos',''])
    return
def criaPlanilhaIndiCrescimento(IndiCrescimento):
    wbsaida.create_sheet('IndiCrescimento')
    IndiCrescimento = wbsaida['IndiCrescimento']
    IndiCrescimento.append(['CAGR Receitas 5 anos', 'CAGR Lucros 5 anos'])
    return
def gravaIndiRentabilidade(wsIndiRentabilidade,teste,coluna,valor):
    print(" teste  "  + valor)
    wsIndiRentabilidade.cell(row=teste, column=coluna, value=valor)
    return

def processamento():
    contador = 1

    for row in tabela.tbody.find_all('tr'):  # a tag <tr>
        # buscando todas as colunas de cada linha
        contador = contador + 1
        te = 1
        print("contator  " + str(contador))
        columns = row.find_all('td')  # a tag <td>
        ida = "teste"  # "#columns[0].text.strip(' ')
        # Grava ROE
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 1, (columns[0].text.strip(' ')))
        # Grava ROIC
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 2, (columns[15].text.strip(' ')))

        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 3, (columns[3].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 4, (columns[4].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 5, (columns[5].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 6, (columns[6].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 7, (columns[7].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 8, (columns[8].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 9, (columns[9].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 10, (columns[10].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 11, (columns[11].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 12, (columns[12].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 13, (columns[13].text.strip(' ')))
        gravaIndiRentabilidade(wsIndiRentabilidade, contador, 14, (columns[14].text.strip(' ')))

        # wsIndiRentabilidade.cell(row=teste, column=1, value=columns[0].text.strip(' '))
        # wsIndiRentabilidade.cell(row=teste, column=2, value=columns[1].text.strip(' '))
        # wsIndiRentabilidade.cell(row=teste, column=3, value=columns[2].text.strip(' '))
        # wsIndiRentabilidade.cell(row=teste, column=4, value=columns[3].text.strip(' '))
        # wsIndiRentabilidade.cell(row=teste, column=5, value=columns[4].text.strip(' '))
        if (columns != []):
            papel = columns[0].text.strip(' ')
            segmento = columns[1].text.strip(' ')
            cotacao = columns[2].text.strip(' ')
            ffo_yield = columns[3].text.strip(' ')
            dividend_yield = columns[4].text.strip(' ')
            p_vp = columns[5].text.strip(' ')
            valor_de_mercado = columns[6].text.strip(' ')
            liquidez = columns[7].text.strip(' ')
            qtd_de_imoveis = columns[8].text.strip(' ')
            preco_do_m2 = columns[9].text.strip(' ')
            aluguel_por_m2 = columns[10].text.strip(' ')
            cap_rate = columns[11].text.strip(' ')
            vacancia_media = columns[13].text.strip(' ')
            print("papel    " + papel, 'vacancia_media    ' + vacancia_media)
            # concatendo o Data Frame criado com as colunas que foram atribuídas a suas colunas
        # wsIndiRentabilidade.cell(row=imposto, column=1, value=papel)
    # wsIndiRentabilidade.cell(row=imposto, column=1, value=porcentagem).number_format = '0%'  # Formato de porcentagem
    return

    return

def gravaCelulaIndiRentabilidade(imposto,wsIndiRentabilidade,te1):

    #values = ws.cell(row=imposto, column=1).value
   # var1 = 'A' + str(imposto)
    nota = imposto + te1
    porcentagem = nota / 100
    #celula[var1] = porcentagem

    #porcentagem_cell = celula[var1] # ws.cell(row=i, column=3, value=porcentagem) b
    #porcentagem_cell =  celula.cell(row=imposto, column=1, value=porcentagem)
    # Definir o formato de porcentagem
    #porcentagem_cell.number_format = '0%'  # Formato de porcentagem
    teste = 1
    for row in tabela.tbody.find_all('tr'):  # a tag <tr>
        # buscando todas as colunas de cada linha
        teste = teste + 1
        te = 1
        print("contator  " + str(teste))
        columns = row.find_all('td')  # a tag <td>
        ida = "teste" #"#columns[0].text.strip(' ')
        #virtual(wsIndiRentabilidade,teste,1,(columns[0].text.strip(' ')))
        #virtual(wsIndiRentabilidade, teste, 2, (columns[1].text.strip(' ')))
        #virtual(wsIndiRentabilidade, teste, 3, (columns[2].text.strip(' ')))
        #virtual(wsIndiRentabilidade, teste, 4, (columns[3].text.strip(' ')))
        #wsIndiRentabilidade.cell(row=teste, column=1, value=columns[0].text.strip(' '))
        #wsIndiRentabilidade.cell(row=teste, column=2, value=columns[1].text.strip(' '))
        #wsIndiRentabilidade.cell(row=teste, column=3, value=columns[2].text.strip(' '))
        #wsIndiRentabilidade.cell(row=teste, column=4, value=columns[3].text.strip(' '))
        #wsIndiRentabilidade.cell(row=teste, column=5, value=columns[4].text.strip(' '))
        if (columns != []):
            papel = columns[0].text.strip(' ')
            segmento = columns[1].text.strip(' ')
            cotacao = columns[2].text.strip(' ')
            ffo_yield = columns[3].text.strip(' ')
            dividend_yield = columns[4].text.strip(' ')
            p_vp = columns[5].text.strip(' ')
            valor_de_mercado = columns[6].text.strip(' ')
            liquidez = columns[7].text.strip(' ')
            qtd_de_imoveis = columns[8].text.strip(' ')
            preco_do_m2 = columns[9].text.strip(' ')
            aluguel_por_m2 = columns[10].text.strip(' ')
            cap_rate = columns[11].text.strip(' ')
            vacancia_media = columns[13].text.strip(' ')
            print("papel    " + papel, 'vacancia_media    ' + vacancia_media)
            # concatendo o Data Frame criado com as colunas que foram atribuídas a suas colunas
           # wsIndiRentabilidade.cell(row=imposto, column=1, value=papel)
   # wsIndiRentabilidade.cell(row=imposto, column=1, value=porcentagem).number_format = '0%'  # Formato de porcentagem
    return

criaPlanilaIndValuation(wbsaida)
criaPlanilhaIndEndividamento(wbsaida)
criaPlanilhaIndiEficiência(wbsaida)
criaPlanilhaIndRentabilidade(wbsaida)
criaPlanilhaIndiCrescimento(wbsaida)

# ler tabela fundamentus


wsIndiRentabilidade = wbsaida['IndiRentabilidade']
processamento()
#gravaCelulaIndiRentabilidade(1,wsIndiRentabilidade,1)
#linha = 1
#while linha < 10:
#    linha = linha + 1
#    wsIndiRentabilidade = wbsaida['IndiRentabilidade']
#    wsIndValuation = wbsaida['IndValuation']
 #   wsIndEndividamento = wbsaida['Endividamento']
  #  wsIndiEficiência = wbsaida['IndiEficiência']
   # wsIndiCrescimento = wbsaida['IndiCrescimento']

 #   gravaCelulaIndiRentabilidade(linha,wsIndiRentabilidade,1)

  #  print(linha)

wbsaida.save("exemplo2.xlsx")
